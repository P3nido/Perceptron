import random #Utilizei para gerar numeros aleatorios nos pesos iniciais e no tetha, para cada treinamento
import time #Utilizei para dar um delay em cada plotagem de grafico, pois como rodei um for 5 vezes ele estava dando problema ao exibir os graficos
import atexit #Utilizei para garantir que o arquivo txt que armazena a saida de prints no meu codigo seja fechado corretamente, apenas para log
import sys #Utilizei para copiar a saida dos prints do arquivo e redirecionar para um txt, para ter um log de todo o meu codigo
from pathlib import Path #Utilizei apenas para criar o caminho do meu arquivo de log saida_executcao.txt
import matplotlib.pyplot as plt #biblioteca para plotar os graficos 
import pandas as pd #biblioteca para organizar os resultados que preenchem a planilha
from openpyxl import load_workbook #biblioteca para escrever os resultados no arquivo xlsx mantendo o layout
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay, accuracy_score, recall_score, precision_score #biblioteca para matriz de confusao e metricas

'''Classe e funcao para redirecionar a saída do print para um arquivo de log, mantendo a saída no console.'''
class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for stream in self.streams:
            stream.write(data)

    def flush(self):
        for stream in self.streams:
            stream.flush()

arquivo_saida = Path(__file__).with_name("saida_execucao.txt")
_stdout_original = sys.stdout
_arquivo_log = open(arquivo_saida, "w", encoding="utf-8")
sys.stdout = Tee(_stdout_original, _arquivo_log)

def _finalizar_log_saida():
    sys.stdout = _stdout_original
    _arquivo_log.close()

atexit.register(_finalizar_log_saida)

def preencher_planilha_questao_2(caminho_planilha, resultados_df):
    if not caminho_planilha.exists():
        print(f"Planilha nao encontrada: {caminho_planilha}")
        return

    workbook = load_workbook(caminho_planilha)
    planilha = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    linha_inicial = 11

    for idx, (_, resultado) in enumerate(resultados_df.iterrows()):
        linha_excel = linha_inicial + idx
        planilha.cell(row=linha_excel, column=7, value=f"{int(resultado['treinamento'])}° Treinamento")
        planilha.cell(row=linha_excel, column=9, value=float(resultado["w0_inicial"]))
        planilha.cell(row=linha_excel, column=10, value=float(resultado["w1_inicial"]))
        planilha.cell(row=linha_excel, column=11, value=float(resultado["w2_inicial"]))
        planilha.cell(row=linha_excel, column=12, value=float(resultado["w3_inicial"]))
        planilha.cell(row=linha_excel, column=13, value=float(resultado["w0_final"]))
        planilha.cell(row=linha_excel, column=14, value=float(resultado["w1_final"]))
        planilha.cell(row=linha_excel, column=15, value=float(resultado["w2_final"]))
        planilha.cell(row=linha_excel, column=16, value=float(resultado["w3_final"]))
        planilha.cell(row=linha_excel, column=17, value=int(resultado["numero_epocas"]))

    workbook.save(caminho_planilha)
    print(f"Planilha atualizada: {caminho_planilha.name}")

def preencher_planilha_questao_4(caminho_planilha, resultados_validacao_df):
    if not caminho_planilha.exists():
        print(f"Planilha nao encontrada: {caminho_planilha}")
        return

    workbook = load_workbook(caminho_planilha)
    planilha = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.active
    linha_inicial = 7
    coluna_inicial = 11  # K = y(t1)

    for linha in range(linha_inicial, linha_inicial + 10):
        for coluna in range(coluna_inicial, coluna_inicial + 5):
            planilha.cell(row=linha, column=coluna, value=None)

    for idx, (_, resultado) in enumerate(resultados_validacao_df.iterrows()):
        linha_excel = linha_inicial + idx
        for treino in range(1, 6):
            coluna_excel = coluna_inicial + (treino - 1)
            coluna_df = f"y_t{treino}"
            valor = resultado.get(coluna_df)
            if pd.notna(valor):
                planilha.cell(row=linha_excel, column=coluna_excel, value=float(valor))

    workbook.save(caminho_planilha)
    print(f"Planilha atualizada: {caminho_planilha.name}")

''' Fase de treinamento do Perceptron com os dados fornecidos'''
dados_treinamento = [
    [-0.6508, 0.1097, 4.0009, -1.0], [-1.4492, 0.8896, 4.4005, -1.0],
    [2.0850, 0.6876, 12.0710, -1.0], [0.2626, 1.1476, 7.7985, 1.0],
    [0.6418, 1.0234, 7.0427, 1.0], [0.2569, 0.6730, 8.3265, -1.0],
    [1.1155, 0.6043, 7.4446, 1.0], [0.0914, 0.3399, 7.0677, -1.0],
    [0.0121, 0.5256, 4.6316, 1.0], [-0.0429, 0.4660, 5.4323, 1.0],
    [0.4340, 0.6870, 8.2287, -1.0], [0.2735, 1.0287, 7.1934, 1.0],
    [0.4839, 0.4851, 7.4850, -1.0], [0.4089, -0.1267, 5.5019, -1.0],
    [1.4391, 0.1614, 8.5843, -1.0], [-0.9115, -0.1973, 2.1962, -1.0],
    [0.3654, 1.0475, 7.4858, 1.0], [0.2144, 0.7515, 7.1699, 1.0],
    [0.2013, 1.0014, 6.5489, 1.0], [0.6483, 0.2183, 5.8991, 1.0],
    [-0.1147, 0.2242, 7.2435, -1.0], [-0.7970, 0.8795, 3.8762, 1.0],
    [-1.0625, 0.6366, 2.4707, 1.0], [0.5307, 0.1285, 5.6883, 1.0],
    [-1.2200, 0.7777, 1.7252, 1.0], [0.3957, 0.1076, 5.6623, -1.0],
    [-0.1013, 0.5989, 7.1812, -1.0], [2.4482, 0.9455, 11.2095, 1.0],
    [2.0149, 0.6192, 10.9263, -1.0], [0.2012, 0.2611, 5.4631, 1.0]
]

coeficiente_de_aprendizado = 0.25 
caminho_planilha_q2 = Path(__file__).with_name("questao-2.xlsx")
caminho_planilha_q4 = Path(__file__).with_name("questao-4.xlsx")
pasta_graficos = Path(__file__).with_name("graficos")
pasta_graficos.mkdir(exist_ok=True)
resultados_treinamentos = []
resultados_validacao_df = pd.DataFrame({"amostra": list(range(1, 11))})
resultados_matriz_confusao = []
resultados_metricas = []

tabela_2 = [
    [-0.3665, 0.0620, 5.9891], [-0.7842, 1.1267, 5.5912],
    [0.3012, 0.5611, 5.8234], [0.7757, 1.0648, 8.0677],
    [0.1570, 0.8028, 6.3040], [-0.7014, 1.0316, 3.6005],
    [0.3748, 0.1536, 6.1537], [-0.6920, 0.9404, 4.4058],
    [-1.3970, 0.7141, 4.9263], [-1.8842, -0.2805, 1.2548]
]

# Vetor real de validacao informado pelo usuario
y_real_validacao = [-1.0, 1.0, 1.0, 1.0, 1.0, 1.0, -1.0, 1.0, -1.0, -1.0]

'''5 treinamentos (for de 0 a 5)'''
for sessao in range(0, 5):
    time.sleep(2)
    
    pesos = [random.uniform(-1, 1) for _ in range(3)] 
    tetha = random.uniform(-1, 1) 
    pesos_iniciais = pesos.copy()
    tetha_inicial = tetha
    
    print(f"Treinamento de numero: {sessao + 1}")
    print(f"Vetor de Pesos Iniciais: w0(theta)={tetha:.5f}, w1={pesos[0]:.5f}, w2={pesos[1]:.5f}, w3={pesos[2]:.5f}")
    
    n_epoca_sessao = 0
    historico_erros = []
    historico_rmse = []
    
    while n_epoca_sessao < 1000:
        '''variavel para contar os erros e RMSE da epoca para plotagem da evolução'''
        erros_vistos = 0 
        soma_erros_quadrados = 0.0 
        
        for linha in dados_treinamento:
            u_temp = sum(xi * wi for xi, wi in zip(linha[0:3], pesos)) + (tetha * -1)
            y_temp = 1.0 if u_temp >= 0 else -1.0
            d_real = linha[3]
            
            if y_temp != d_real:
                erros_vistos += 1
            
            soma_erros_quadrados += (d_real - y_temp) ** 2
                
        historico_erros.append(erros_vistos)
        
        rmse_epoca = (soma_erros_quadrados / len(dados_treinamento)) ** 0.5
        historico_rmse.append(rmse_epoca)

        erro_na_epoca = False
        for linha in dados_treinamento:
            x = linha[0:3]
            d = linha[3]
            
            u = sum(xi * wi for xi, wi in zip(x, pesos)) + (tetha * -1)
            y = 1.0 if u >= 0 else -1.0
            
            if y != d:
                erro_na_epoca = True
                for i in range(3):
                    pesos[i] = pesos[i] + coeficiente_de_aprendizado * (d - y) * x[i]
                tetha = tetha + coeficiente_de_aprendizado * (d - y) * (-1)
        
        n_epoca_sessao += 1
        
        if not erro_na_epoca:
            break

    print(f"Numero de epocas: {n_epoca_sessao}")
    print(f"Vetor de Pesos Finais: w0(theta)={tetha:.5f}, w1={pesos[0]:.5f}, w2={pesos[1]:.5f}, w3={pesos[2]:.5f}")
    print("-" * 50)

    resultado_atual = {
        "treinamento": sessao + 1,
        "w0_inicial": round(tetha_inicial, 5),
        "w1_inicial": round(pesos_iniciais[0], 5),
        "w2_inicial": round(pesos_iniciais[1], 5),
        "w3_inicial": round(pesos_iniciais[2], 5),
        "w0_final": round(tetha, 5),
        "w1_final": round(pesos[0], 5),
        "w2_final": round(pesos[1], 5),
        "w3_final": round(pesos[2], 5),
        "numero_epocas": n_epoca_sessao,
    }
    resultados_treinamentos.append(resultado_atual)
    resultados_treinamentos_df = pd.DataFrame(resultados_treinamentos)
    preencher_planilha_questao_2(caminho_planilha_q2, resultados_treinamentos_df)

    resultados_validacao_sessao = []
    for x_teste in tabela_2:
        u_validacao = sum(xi * wi for xi, wi in zip(x_teste, pesos)) + (tetha * -1)
        y_validacao = 1.0 if u_validacao >= 0 else -1.0
        resultados_validacao_sessao.append(y_validacao)

    matriz_confusao = confusion_matrix(
        y_real_validacao,
        resultados_validacao_sessao,
        labels=[-1.0, 1.0]
    )

    verdadeiro_negativo, falso_positivo, falso_negativo, verdadeiro_positivo = matriz_confusao.ravel()
    numero_de_acertos = sum(
        1 for valor_real, valor_previsto in zip(y_real_validacao, resultados_validacao_sessao)
        if valor_real == valor_previsto
    )
    numero_de_erros = len(y_real_validacao) - numero_de_acertos

    acuracia = accuracy_score(y_real_validacao, resultados_validacao_sessao)
    sensibilidade = recall_score(y_real_validacao, resultados_validacao_sessao, pos_label=1.0, zero_division=0)
    especificidade = (
        verdadeiro_negativo / (verdadeiro_negativo + falso_positivo)
        if (verdadeiro_negativo + falso_positivo) > 0
        else 0.0
    )
    precisao = precision_score(y_real_validacao, resultados_validacao_sessao, pos_label=1.0, zero_division=0)

    resultados_metricas.append({
        "Treinamento": sessao + 1,
        "Verdadeiro Negativo": int(verdadeiro_negativo),
        "Falso Positivo": int(falso_positivo),
        "Falso Negativo": int(falso_negativo),
        "Verdadeiro Positivo": int(verdadeiro_positivo),
        "Numero de Acertos": int(numero_de_acertos),
        "Numero de Erros": int(numero_de_erros),
        "Acuracia": round(acuracia, 4),
        "Sensibilidade": round(sensibilidade, 4),
        "Especificidade": round(especificidade, 4),
        "Precisao": round(precisao, 4),
    })

    resultados_matriz_confusao.append({
        "treinamento": sessao + 1,
        "matriz": matriz_confusao
    })

    print(f"Matriz de confusao - Treinamento {sessao + 1}:")
    print(matriz_confusao)

    fig_cm, ax_cm = plt.subplots(figsize=(5, 4))
    disp_cm = ConfusionMatrixDisplay(
        confusion_matrix=matriz_confusao,
        display_labels=["P1 (-1)", "P2 (+1)"]
    )
    disp_cm.plot(ax=ax_cm, cmap="Blues", colorbar=False, values_format="d")
    ax_cm.set_title(f"Matriz de Confusao - Treinamento {sessao + 1}")
    caminho_matriz = pasta_graficos / f"matriz_confusao_treinamento_{sessao + 1}.png"
    fig_cm.tight_layout()
    fig_cm.savefig(caminho_matriz, dpi=150)
    print(f"Matriz de confusao salva em: {caminho_matriz.name}")
    plt.close(fig_cm)

    print(f"Numero de Acertos: {numero_de_acertos} | Numero de Erros: {numero_de_erros}")
    print(
        f"Acuracia: {acuracia:.4f} | Sensibilidade: {sensibilidade:.4f} | "
        f"Especificidade: {especificidade:.4f} | Precisao: {precisao:.4f}"
    )

    resultados_validacao_df[f"y_t{sessao + 1}"] = resultados_validacao_sessao
    preencher_planilha_questao_4(caminho_planilha_q4, resultados_validacao_df)


    ''' plotando a evolução de aprendizado para cada treinamento '''
    plt.figure(figsize=(8, 4))
    plt.plot(range(1, n_epoca_sessao + 1), historico_erros, color='blue')
    plt.title(f'Evolucao do Erro - Treinamento {sessao + 1}')
    plt.xlabel('Epoca')
    plt.ylabel('Quantidade de Erros')
    plt.grid(True)
    caminho_grafico = pasta_graficos / f"grafico_treinamento_{sessao + 1}.png"
    plt.tight_layout()
    plt.savefig(caminho_grafico, dpi=150)
    print(f"Grafico salvo em: {caminho_grafico.name}")
    plt.close()

    ''' plotando a evolução do RMSE para cada treinamento '''
    plt.figure(figsize=(8, 4))
    plt.plot(range(1, n_epoca_sessao + 1), historico_rmse, color='red')
    plt.title(f'Evolucao do RMSE - Treinamento {sessao + 1}')
    plt.xlabel('Epoca')
    plt.ylabel('RMSE (Raiz do Erro Quadratico Medio)')
    plt.grid(True)
    caminho_grafico_rmse = pasta_graficos / f"grafico_rmse_treinamento_{sessao + 1}.png"
    plt.tight_layout()
    plt.savefig(caminho_grafico_rmse, dpi=150)
    print(f"Grafico RMSE salvo em: {caminho_grafico_rmse.name}")
    plt.close()

print("\nAPRENDIZADO FINALIZADO!")
time.sleep(2)

print("\nRESUMO DAS MATRIZES DE CONFUSAO (5 TREINAMENTOS):")
for resultado in resultados_matriz_confusao:
    print(f"Treinamento {resultado['treinamento']}:")
    print(resultado["matriz"])
    print("-" * 40)

print("\nTABELA DE METRICAS (PANDAS):")
tabela_metricas_df = pd.DataFrame(resultados_metricas)
print(tabela_metricas_df.to_string(index=False))

caminho_tabela_metricas = pasta_graficos / "tabela_metricas_treinamentos.csv"
tabela_metricas_df.to_csv(caminho_tabela_metricas, index=False, sep=";")
print(f"Tabela de metricas salva em: {caminho_tabela_metricas.name}")

'''Fase de validação final do Perceptron usando os pesos do ultimo treinamento.'''
print("-" * 40)
print("RESULTADOS DA VALIDAÇÃO (TREINAMENTO FINAL):")
for i, x_teste in enumerate(tabela_2, 1):
    u_op = sum(xi * wi for xi, wi in zip(x_teste, pesos)) + (tetha * -1)
    y_op = 1.0 if u_op >= 0 else -1.0
    classe = "P2" if y_op == 1 else "P1"
    print(f"Amostra {i}: y = {y_op} -> Classe {classe}")